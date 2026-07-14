import os
import io
import json
import math
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaInMemoryUpload

# ── SETTINGS ──────────────────────────────────────────────────────────────────
LOOKBACK   = "5y"
RSI_PERIOD = 14
RSI_MA     = 14

SECTORS = {
    "DJI":  "Dow Jones Index",
    "SPY":  "S&P 500 Index",
    "QQQ":  "Nasdaq 100 Index",
    "IWM":  "Russell 2000 Index",
    "XLK":  "Technology",
    "XLF":  "Financials",
    "XLV":  "Health Care",
    "XLY":  "Consumer Discretionary",
    "XLC":  "Communication Services",
    "XLI":  "Industrials",
    "XLP":  "Consumer Staples",
    "XLE":  "Energy",
    "XLU":  "Utilities",
    "XLRE": "Real Estate",
    "XLB":  "Materials",
}

WATCHLIST_SEED = [
    "AMD","VRT","PLTR","FUTU","SHOP","DELL","CRDO","ANET","HOOD","WDC",
    "CCJ","KTOS","FTNT","INOD","CSCO","IBIT","META","APP","MSFT","TSLA",
    "AXP","AVGO","GE","JPM","CLS","TSM","AAPL","GOOGL","STX","AMZN",
    "MU","NVDA","ETHA","SOFI","CDE","IREN","AA","ADI","CCL","HL",
    "AMAT","LRCX","APH","EQT","NEM","CAT","FCX","RTX","GLW","COHR",
    "DRAM","INTC","SMH","CEG","NBIS","TER"
]

FINANCIALS_SEED = [
    "JPM","BAC","WFC","GS","MS","C","AXP","BLK","SCHW","COF"
]

# ── Analysis helpers ──────────────────────────────────────────────────────────
def calc_ema(series, span):
    return series.ewm(span=span, adjust=False).mean()

def calc_rsi(close, period=14):
    delta    = close.diff()
    gain     = delta.clip(lower=0)
    loss     = -delta.clip(upper=0)
    avg_gain = gain.ewm(com=period - 1, adjust=False).mean()
    avg_loss = loss.ewm(com=period - 1, adjust=False).mean()
    rs       = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def get_last_crossover(series_a, series_b, index):
    last_date = None
    last_type = None
    for i in range(1, len(series_a)):
        prev = float(series_a.iloc[i-1]) - float(series_b.iloc[i-1])
        curr = float(series_a.iloc[i])   - float(series_b.iloc[i])
        if prev <= 0 and curr > 0:
            last_date = index[i].strftime("%Y-%m-%d")
            last_type = "▲ Bullish"
        elif prev >= 0 and curr < 0:
            last_date = index[i].strftime("%Y-%m-%d")
            last_type = "▼ Bearish"
    return last_date or "N/A", last_type or "N/A"

def get_iv_30d(ticker):
    """Calculate approximate 30-day IV from ATM options closest to 30 DTE."""
    try:
        t = yf.Ticker(ticker)
        exps = t.options
        if not exps:
            return "N/A"

        today      = date.today()
        target     = today + timedelta(days=30)
        best_exp   = min(exps, key=lambda e: abs((date.fromisoformat(e) - target).days))
        dte        = (date.fromisoformat(best_exp) - today).days
        if dte <= 0:
            return "N/A"

        hist = t.history(period="5d")
        if hist.empty:
            return "N/A"
        spot = float(hist["Close"].iloc[-1])

        chain = t.option_chain(best_exp)
        calls = chain.calls
        puts  = chain.puts

        # Find ATM call and put (closest strike to spot)
        if calls.empty and puts.empty:
            return "N/A"

        ivs = []
        for df in [calls, puts]:
            if df.empty:
                continue
            df = df.copy()
            df["dist"] = (df["strike"] - spot).abs()
            atm = df.nsmallest(2, "dist")
            for _, row in atm.iterrows():
                iv = row.get("impliedVolatility",
